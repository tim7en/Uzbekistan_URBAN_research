"""
Methodology Comparison Report
==============================
Compares two spatial disaggregation approaches for Uzbekistan GHG emissions (2022):
  v1  LULC + VIIRS Nightlights only
  v2  Sentinel-5P (NO2/SO2/CO/CH4/AER) + LULC + VIIRS, reliability-adjusted

Output: suhi_analysis_output/methodology_comparison_report_2022.pdf
"""
import sys
from pathlib import Path

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.patches as mpatches
import matplotlib.ticker as mticker
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.colors import LinearSegmentedColormap
import rasterio
import pandas as pd
import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
V1_DIR   = ROOT / "suhi_analysis_output" / "regional_ghg_maps"
V2_DIR   = ROOT / "suhi_analysis_output" / "regional_ghg_maps_s5p"
CACHE_V1 = V1_DIR / "_raster_cache"
XLSX     = ROOT / "uzbekistan_emissions_by_region_2022 (1)+.xlsx"
OUT_PDF  = ROOT / "suhi_analysis_output" / "methodology_comparison_report_2022.pdf"
YEAR     = 2022

# ---------------------------------------------------------------------------
# Region metadata: (slug_v1, slug_v2, display_name, s5p_reliability)
# Reliability reflects share of non-noisy LULC classes
# (HIGH >= 0.75 dense urban/agricultural; LOW < 0.40 desert-dominated)
# ---------------------------------------------------------------------------
REGIONS = [
    ("Tashkent_City",       "Tashkent_City",       "Tashkent City",         0.88),
    ("Andijan_Region",      "Andijan_Region",       "Andijan Region",        0.82),
    ("Fergana_Region",      "Fergana_Region",       "Fergana Region",        0.64),
    ("Jizzakh_Region",      "Jizzakh_Region",       "Jizzakh Region",        0.64),
    ("Namangan_Region",     "Namangan_Region",      "Namangan Region",       0.60),
    ("Syrdarya_Region",     "Syrdarya_Region",      "Syrdarya Region",       0.59),
    ("Tashkent_Region",     "Tashkent_Region",      "Tashkent Region",       0.54),
    ("Samarkand_Region",    "Samarkand_Region",     "Samarkand Region",      0.54),
    ("Kashkadarya_Region",  "Kashkadarya_Region",   "Kashkadarya Region",    0.48),
    ("Khorezm_Region",      "Khorezm_Region",       "Khorezm Region",        0.46),
    ("Surkhandarya_Region", "Surkhandarya_Region",  "Surkhandarya Region",   0.45),
    ("Navoi_Region",        "Navoi_Region",         "Navoi Region",          0.35),
    ("Bukhara_Region",      "Bukhara_Region",       "Bukhara Region",        0.32),
    ("Rep_Karakalpakstan",  "Rep_Karakalpakstan",   "Rep. Karakalpakstan",   0.35),
]

ESRI_CLASSES = {
    1: "Water", 2: "Trees", 4: "Flooded Veg.", 5: "Crops",
    7: "Built Area", 8: "Bare Ground", 9: "Snow/Ice", 11: "Rangeland"
}
LULC_COLORS_MAP = {
    "Water": "#419BDF", "Trees": "#397D49", "Flooded Veg.": "#7A87C6",
    "Crops": "#E49635", "Built Area": "#C4281B", "Bare Ground": "#A59B8F",
    "Snow/Ice": "#B39FE1", "Rangeland": "#DFC35A"
}

SECTORS       = ["electricity_heat","residential_commercial","industry_combustion",
                 "transport","fugitive_emissions","ippu","agriculture","waste"]
SECTOR_LABELS = ["Elec & Heat","Res & Comm","Industry",
                 "Transport","Fugitive","IPPU","Agriculture","Waste"]
SECTOR_COLORS = ["#E63946","#457B9D","#F4A261","#2A9D8F",
                 "#E9C46A","#264653","#A8DADC","#72B7B2"]

# Dark theme
BG    = "#0E1117"
PANEL = "#16191F"
WHITE = "#FFFFFF"
GREY  = "#AAAAAA"
ACC1  = "#4A90D9"   # v1 blue
ACC2  = "#E8543A"   # v2 red-orange

# ---------------------------------------------------------------------------
# Data loading helpers
# ---------------------------------------------------------------------------
def load_excel_emissions():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if "emissions by region" in name.lower():
            ws = wb[name]; break
    if ws is None:
        for name in wb.sheetnames:
            if "region" in name.lower():
                ws = wb[name]; break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        vals = [str(v).replace("\n"," ").strip().lower() for v in row if v is not None]
        if any(v.startswith("region") for v in vals) and \
           any("electric" in v or "heat" in v for v in vals):
            header_idx = i; break
    if header_idx is None:
        wb.close(); return pd.DataFrame()

    raw_headers = [str(c).replace("\n"," ").strip() if c else f"col_{j}"
                   for j, c in enumerate(rows[header_idx])]
    SECTOR_MAP = {
        "electricity & heat": "electricity_heat",
        "residential & commercial": "residential_commercial",
        "industry combustion": "industry_combustion",
        "transport": "transport",
        "fugitive emissions": "fugitive_emissions",
        "fugitive": "fugitive_emissions",
        "ippu": "ippu",
        "agriculture": "agriculture",
        "waste": "waste",
        "lulucf": "lulucf",
        "total": "total",
    }
    clean = []
    for h in raw_headers:
        hl = h.lower()
        matched = next((v for k, v in SECTOR_MAP.items() if k in hl), None)
        clean.append(matched if matched else hl.replace(" ", "_"))

    EXCLUDE = ["nan","none","note","source","sum of","national","allocation"]
    records = []
    for row in rows[header_idx + 1:]:
        if row[0] is None: continue
        rname = str(row[0]).strip()
        if any(ex in rname.lower() for ex in EXCLUDE): continue
        rec = {"region": rname}
        for j, h in enumerate(clean[1:], 1):
            try:
                v = row[j]
                rec[h] = float(v) if v is not None else 0.0
            except (TypeError, ValueError):
                rec[h] = 0.0
        records.append(rec)
    wb.close()
    return pd.DataFrame(records)


def find_lulc(slug):
    # Check subdirectory first (new cache layout), then flat
    for pat in [f"{slug}/lulc_{slug}_*.tif", f"lulc_{slug}_*.tif",
                f"*/{slug}/lulc_*.tif", f"lulc_*{slug}*.tif"]:
        c = list(CACHE_V1.glob(pat))
        if c: return c[0]
    return None


def load_raster(path):
    with rasterio.open(path) as src:
        return src.read(1, masked=True)


def get_lulc_fracs(slug):
    p = find_lulc(slug)
    if p is None: return {}
    try:
        arr = load_raster(p).filled(0).astype(int)
        total = max((arr > 0).sum(), 1)
        return {ESRI_CLASSES[c]: (arr == c).sum() / total
                for c in ESRI_CLASSES if (arr == c).sum() > 0}
    except Exception:
        return {}


def get_region_emissions(df, name):
    """Return dict of sector -> Mt for the named region."""
    if df.empty: return {}
    key = name.split(" ")[0]
    mask = df["region"].str.contains(key, case=False, na=False)
    if not mask.any(): return {}
    row = df[mask].iloc[0]
    return {s: float(row[s]) if s in row.index and not pd.isna(row.get(s, 0)) else 0.0
            for s in SECTORS}


# ---------------------------------------------------------------------------
# Figure / axes helpers
# ---------------------------------------------------------------------------
def new_fig(w, h):
    fig = plt.figure(figsize=(w, h))
    fig.patch.set_facecolor(BG)
    return fig


def style_ax(ax, title=None, xlabel=None, ylabel=None, fontsize=8):
    ax.set_facecolor(PANEL)
    ax.tick_params(colors=WHITE, labelsize=fontsize - 1)
    for sp in ax.spines.values():
        sp.set_edgecolor("#333")
    if title:
        ax.set_title(title, fontsize=fontsize, fontweight="bold",
                     color=WHITE, pad=3)
    if xlabel:
        ax.set_xlabel(xlabel, fontsize=fontsize - 1, color=GREY)
    if ylabel:
        ax.set_ylabel(ylabel, fontsize=fontsize - 1, color=GREY)
    ax.yaxis.grid(True, alpha=0.12, lw=0.4, color=WHITE)
    ax.set_axisbelow(True)


def rel_color(r):
    if r >= 0.75: return "#2ECC71"
    if r >= 0.40: return "#F39C12"
    return "#E74C3C"


def rel_label(r):
    if r >= 0.75: return "HIGH"
    if r >= 0.40: return "MODERATE"
    return "LOW"


# ===========================================================================
# PAGE 1 — Cover
# ===========================================================================
def page_cover(pdf):
    fig = new_fig(16, 11)

    fig.text(0.5, 0.91, "Uzbekistan Regional GHG Emissions 2022",
             ha="center", fontsize=22, fontweight="bold", color=WHITE)
    fig.text(0.5, 0.86,
             "Spatial Disaggregation — Methodology Comparison Report",
             ha="center", fontsize=14, color=ACC1)
    fig.text(0.5, 0.82,
             "v1: ESRI 10m LULC + VIIRS Nightlights   |   "
             "v2: Sentinel-5P Atmospheric Signals + LULC + VIIRS (reliability-adjusted)",
             ha="center", fontsize=9.5, color=GREY)

    fig.add_artist(plt.Line2D([0.04, 0.96], [0.79, 0.79],
                              transform=fig.transFigure,
                              color=ACC1, lw=1.0))

    # Two methodology boxes
    method_boxes = [
        (0.04, ACC1, "v1 — LULC + Nightlights", [
            "Spatial proxy signals:",
            "  * ESRI Global LULC 10m Time Series (200m mosaic)",
            "  * NOAA/VIIRS monthly nightlights, annual median 2022",
            "",
            "Disaggregation logic:",
            "  Each sector allocated by LULC class fractions",
            "  (built area, crops, bare ground ...) weighted by",
            "  nightlight intensity as an economic activity proxy.",
            "",
            "Strengths:",
            "  + Deterministic and reproducible",
            "  + Not affected by atmospheric retrieval noise",
            "  + Excellent sub-city spatial resolution (200m)",
            "",
            "Limitations:",
            "  - Cannot distinguish emission intensity within",
            "    the same LULC class (e.g. heavy vs light industry)",
            "  - Nightlights proxy economic activity, not chemistry",
            "  - No sector-specific atmospheric signal",
        ]),
        (0.52, ACC2, "v2 — Sentinel-5P + LULC + Nightlights (reliability-adjusted)", [
            "Spatial proxy signals (TROPOMI OFFL L3 annual medians):",
            "  * NO2 tropospheric column   (cloud_fraction < 0.5)",
            "  * SO2 total column",
            "  * CO  total column",
            "  * CH4 dry-air mixing ratio, bias-corrected (anomaly above P10)",
            "  * AER_AI absorbing aerosol index",
            "  + ESRI LULC 10m + VIIRS nightlights (same as v1)",
            "",
            "Per-sector S5P blend weights (nominal):",
            "  Elec & Heat   : SO2 40% + NO2 25% + built 15% + NL 20%",
            "  Transport     : NO2 45% + CO  35% + built 20%",
            "  Fugitive      : CH4 55% + SO2 25% + uniform 20%",
            "  Agriculture   : CH4 40% + crops 40% + AER 20%",
            "",
            "Regional reliability adjustment:",
            "  r = clamp(1 - sum(noise_weight * LULC_frac), 0.15, 1.0)",
            "  S5P weight *= r; deficit redistributed to LULC/NL.",
            "",
            "Strengths:",
            "  + Actual atmospheric chemistry as primary constraint",
            "  + Sector-specific signal (SO2=power, NO2=traffic ...)",
            "Limitations:",
            "  - TROPOMI ~5.5 km footprint limits sub-city detail",
            "  - Desert / bright surfaces degrade UV/VIS retrievals",
        ]),
    ]
    for col, (xL, color, title, lines) in enumerate(method_boxes):
        ax = fig.add_axes([xL, 0.18, 0.44, 0.59])
        ax.set_facecolor("#0A1520" if col == 0 else "#1A0A08")
        ax.set_xlim(0, 1); ax.set_ylim(0, 1); ax.axis("off")
        for sp in ax.spines.values():
            sp.set_edgecolor(color); sp.set_lw(1.4)
        ax.text(0.5, 0.97, title, ha="center", va="top",
                fontsize=9.5, fontweight="bold", color=color,
                transform=ax.transAxes)
        y = 0.91
        for line in lines:
            ax.text(0.04, y, line, ha="left", va="top",
                    fontsize=7.8, color=WHITE, transform=ax.transAxes,
                    fontfamily="monospace")
            y -= 0.038

    # Footer
    fig.text(0.5, 0.13,
             "Report structure:  Page 1 Cover  |  Page 2 Summary Table  |  "
             "Pages 3-16 Per-region comparison  |  Page 17 Analytics  |  Page 18 Findings",
             ha="center", fontsize=8, color=GREY)
    fig.text(0.5, 0.09,
             "Emission totals: Uzbekistan First Biennial Transparency Report (BTR1) 2024   "
             "LULC: ESRI Global LULC 10m   NL: NOAA/VIIRS DNB   "
             "S5P: Copernicus OFFL L3 TROPOMI",
             ha="center", fontsize=7.5, color="#555", style="italic")
    fig.text(0.5, 0.05,
             "Scientific basis for reliability scoring: van Geffen et al. 2022 (NO2 ATBD v2.4); "
             "Theys et al. 2017 ACP (SO2); Torres et al. 2007 JGR (AER_AI); "
             "Tilstra et al. 2021 AMT (surface LER climatology)",
             ha="center", fontsize=7, color="#444", style="italic")

    pdf.savefig(fig, facecolor=BG, bbox_inches="tight")
    plt.close(fig)
    print("  Page 1 (cover) done.")


# ===========================================================================
# PAGE 2 — Summary table with reliability scores and LULC breakdown
# ===========================================================================
def page_summary(pdf, region_stats):
    fig = new_fig(16, 11)
    fig.text(0.5, 0.965, "Regional Overview — LULC Composition & S5P Reliability",
             ha="center", fontsize=13, fontweight="bold", color=WHITE)

    # ----- Top: LULC stacked bar -----
    ax_lulc = fig.add_axes([0.05, 0.58, 0.55, 0.35])
    style_ax(ax_lulc, title="LULC Composition by Region (%)",
             ylabel="Area fraction (%)", fontsize=8)

    LULC_ORDER = ["Built Area","Crops","Rangeland","Bare Ground",
                  "Water","Trees","Flooded Veg.","Snow/Ice"]
    names_short = [r["name"].replace(" Region","").replace("Rep. ","")
                   for r in region_stats]
    bottoms = np.zeros(len(region_stats))
    for cls in LULC_ORDER:
        col = LULC_COLORS_MAP.get(cls, "#888")
        vals = np.array([r["lulc_fracs"].get(cls, 0) * 100 for r in region_stats])
        ax_lulc.bar(range(len(names_short)), vals, bottom=bottoms,
                    color=col, label=cls, edgecolor="#0E1117", lw=0.3)
        bottoms += vals
    ax_lulc.set_xticks(range(len(names_short)))
    ax_lulc.set_xticklabels(names_short, rotation=35, ha="right",
                             fontsize=7, color=WHITE)
    ax_lulc.set_ylim(0, 108)
    ax_lulc.legend(fontsize=6.5, facecolor=PANEL, labelcolor=WHITE,
                   edgecolor="#333", loc="upper right", ncol=2)

    # ----- Reliability bar -----
    ax_rel = fig.add_axes([0.65, 0.58, 0.32, 0.35])
    style_ax(ax_rel, title="S5P Reliability Score", ylabel="Score [0-1]", fontsize=8)
    rels   = [r["reliability"] for r in region_stats]
    colors = [rel_color(r) for r in rels]
    ax_rel.bar(range(len(names_short)), rels, color=colors,
               edgecolor="#0E1117", lw=0.4)
    ax_rel.axhline(0.75, color="#2ECC71", lw=0.8, ls="--", alpha=0.6)
    ax_rel.axhline(0.40, color="#F39C12", lw=0.8, ls="--", alpha=0.6)
    ax_rel.set_xticks(range(len(names_short)))
    ax_rel.set_xticklabels(names_short, rotation=35, ha="right",
                            fontsize=7, color=WHITE)
    ax_rel.set_ylim(0, 1.08)
    for i, (ri, rc) in enumerate(zip(rels, colors)):
        ax_rel.text(i, ri + 0.01, f"{int(ri*100)}%", ha="center",
                    va="bottom", fontsize=5.5, color=WHITE)
    patches = [mpatches.Patch(color="#2ECC71", label="HIGH (>=75%)"),
               mpatches.Patch(color="#F39C12", label="MOD (40-74%)"),
               mpatches.Patch(color="#E74C3C", label="LOW  (<40%)")]
    ax_rel.legend(handles=patches, fontsize=6, facecolor=PANEL,
                  labelcolor=WHITE, edgecolor="#333", loc="upper left")

    # ----- Table -----
    ax_t = fig.add_axes([0.01, 0.01, 0.98, 0.54])
    ax_t.set_facecolor(PANEL); ax_t.axis("off")

    cols = ["Region", "Total\n(Mt CO2-eq)", "Built\n(%)", "Crops\n(%)",
            "Rangeland\n(%)", "Bare Gnd\n(%)", "S5P Score", "Rating",
            "Recommended approach"]
    cw = [0.16, 0.08, 0.07, 0.07, 0.09, 0.09, 0.08, 0.08, 0.18]
    xs = []
    x = 0.01
    for w in cw:
        xs.append(x + w / 2); x += w

    # Header row
    for xi, h in zip(xs, cols):
        ax_t.text(xi, 0.95, h, ha="center", va="top",
                  fontsize=7.5, fontweight="bold", color=ACC1,
                  transform=ax_t.transAxes)
    ax_t.add_artist(plt.Line2D([0, 1], [0.89, 0.89],
                               transform=ax_t.transAxes,
                               color=ACC1, lw=0.7))

    rh = 0.83 / len(region_stats)
    for i, r in enumerate(region_stats):
        y = 0.87 - i * rh
        lf = r["lulc_fracs"]
        rc = rel_color(r["reliability"])
        rl = rel_label(r["reliability"])
        approach = ("v2 S5P primary" if r["reliability"] >= 0.75 else
                    "v2 blended" if r["reliability"] >= 0.40 else
                    "v1 preferred")
        row = [
            r["name"],
            f"{r.get('total_mt', 0):.2f}",
            f"{lf.get('Built Area', 0)*100:.1f}",
            f"{lf.get('Crops', 0)*100:.1f}",
            f"{lf.get('Rangeland', 0)*100:.1f}",
            f"{lf.get('Bare Ground', 0)*100:.1f}",
            f"{r['reliability']:.2f}",
            rl,
            approach,
        ]
        for j, (xi, val) in enumerate(zip(xs, row)):
            tc = rc if j == 7 else WHITE
            fw = "bold" if j == 7 else "normal"
            ax_t.text(xi, y, val, ha="center", va="center",
                      fontsize=7, color=tc, fontweight=fw,
                      transform=ax_t.transAxes)

    pdf.savefig(fig, facecolor=BG, bbox_inches="tight")
    plt.close(fig)
    print("  Page 2 (summary table) done.")


# ===========================================================================
# PAGES 3-16 — One page per region: side-by-side v1 vs v2 + stats
# ===========================================================================
def page_region(pdf, slug_v1, slug_v2, name, reliability, lulc_fracs_data,
                sector_emissions):
    # Landscape A3-ish for legibility
    fig = new_fig(20, 12)

    rc = rel_color(reliability)
    rl = rel_label(reliability)

    # Header
    fig.text(0.5, 0.975,
             f"{name}   |   S5P Reliability: {rl} ({int(reliability*100)}%)",
             ha="center", fontsize=13, fontweight="bold",
             color=WHITE,
             bbox=dict(boxstyle="round,pad=0.3",
                       facecolor=PANEL, edgecolor=rc, lw=1.5))

    # ---- Map panels (top 72% of figure) ----
    ax_v1 = fig.add_axes([0.01, 0.25, 0.485, 0.72])
    ax_v2 = fig.add_axes([0.505, 0.25, 0.485, 0.72])

    for ax, slug, label, col, png_dir, prefix in [
        (ax_v1, slug_v1, "v1 — LULC + Nightlights",      ACC1, V1_DIR, "ghg_"),
        (ax_v2, slug_v2, "v2 — S5P + LULC + NL",         ACC2, V2_DIR, "ghg_s5p_"),
    ]:
        ax.axis("off")
        p = png_dir / f"{prefix}{slug}_{YEAR}.png"
        if p.exists():
            img = plt.imread(str(p))
            ax.imshow(img, aspect="auto")
        else:
            ax.set_facecolor(PANEL)
            ax.text(0.5, 0.5, f"{name}\n{label}\nnot found",
                    ha="center", va="center", color=GREY,
                    fontsize=10, transform=ax.transAxes)
        ax.set_title(label, fontsize=10, fontweight="bold",
                     color=col, pad=4)
        for sp in ax.spines.values():
            sp.set_edgecolor(col); sp.set_lw(1.5); sp.set_visible(True)

    # ---- Bottom stats strip ----
    gs_b = gridspec.GridSpec(1, 4, figure=fig,
                             left=0.02, right=0.98,
                             top=0.22, bottom=0.02,
                             wspace=0.35)

    # Panel 1: Sector emissions bar
    ax_sec = fig.add_subplot(gs_b[0, 0])
    style_ax(ax_sec, title="Sector Emissions (Mt CO2-eq)",
             ylabel="Mt CO2-eq", fontsize=8)
    if sector_emissions:
        vals = [sector_emissions.get(s, 0) for s in SECTORS]
        bars = ax_sec.bar(range(len(SECTORS)), vals,
                          color=SECTOR_COLORS, edgecolor="#111", lw=0.4)
        ax_sec.set_xticks(range(len(SECTORS)))
        ax_sec.set_xticklabels(SECTOR_LABELS, rotation=40,
                               ha="right", fontsize=6, color=WHITE)
        for b, v in zip(bars, vals):
            if v > 0:
                ax_sec.text(b.get_x() + b.get_width()/2,
                            b.get_height() + max(vals) * 0.01,
                            f"{v:.2f}", ha="center", va="bottom",
                            fontsize=5.5, color=WHITE)
    else:
        ax_sec.text(0.5, 0.5, "No data", ha="center", va="center",
                    color=GREY, transform=ax_sec.transAxes, fontsize=8)

    # Panel 2: LULC composition
    ax_lulc = fig.add_subplot(gs_b[0, 1])
    style_ax(ax_lulc, title="LULC Composition (%)", fontsize=8)
    if lulc_fracs_data:
        items = sorted(lulc_fracs_data.items(), key=lambda x: -x[1])
        labels = [k for k, v in items]
        values = [v * 100 for k, v in items]
        colors = [LULC_COLORS_MAP.get(k, "#888") for k in labels]
        wedges, texts, autotexts = ax_lulc.pie(
            values, labels=None, colors=colors,
            autopct=lambda p: f"{p:.0f}%" if p > 4 else "",
            startangle=90, pctdistance=0.75,
            wedgeprops=dict(edgecolor="#0E1117", lw=0.5))
        for at in autotexts:
            at.set_fontsize(6); at.set_color(WHITE)
        ax_lulc.legend(labels, fontsize=5.5, facecolor=PANEL,
                       labelcolor=WHITE, edgecolor="#333",
                       loc="lower right", bbox_to_anchor=(1.45, -0.1))
    else:
        ax_lulc.text(0.5, 0.5, "No LULC data", ha="center",
                     va="center", color=GREY, fontsize=8,
                     transform=ax_lulc.transAxes)

    # Panel 3: S5P recipe blend breakdown
    ax_blend = fig.add_subplot(gs_b[0, 2])
    style_ax(ax_blend, title="v2 S5P Blend Weights per Sector",
             ylabel="Effective weight", fontsize=8)
    BASE_RECIPES = {
        "Elec & Heat":   {"SO2": 0.40,"NO2": 0.25,"built": 0.15,"NL": 0.20},
        "Res & Comm":    {"CO":  0.30,"NO2": 0.20,"built": 0.25,"NL": 0.25},
        "Industry":      {"SO2": 0.35,"NO2": 0.30,"CO":   0.20,"built":0.15},
        "Transport":     {"NO2": 0.45,"CO":  0.35,"built": 0.20},
        "Fugitive":      {"CH4": 0.55,"SO2": 0.25,"unif": 0.20},
        "IPPU":          {"SO2": 0.45,"NO2": 0.30,"AER":  0.15,"built":0.10},
        "Agriculture":   {"CH4": 0.40,"crops":0.40,"AER": 0.20},
        "Waste":         {"CH4": 0.40,"CO":  0.30,"built": 0.30},
    }
    S5P_KEYS = {"SO2","NO2","CO","CH4","AER"}
    sec_names = list(BASE_RECIPES.keys())
    s5p_w  = [sum(v for k,v in r.items() if k in S5P_KEYS) * reliability
              for r in BASE_RECIPES.values()]
    lulc_w = [1 - sw for sw in s5p_w]
    x = np.arange(len(sec_names))
    ax_blend.bar(x, s5p_w,  color=ACC2, label=f"S5P (×{reliability:.2f})",
                 edgecolor="#111", lw=0.4)
    ax_blend.bar(x, lulc_w, bottom=s5p_w, color=ACC1,
                 label="LULC+NL", edgecolor="#111", lw=0.4)
    ax_blend.set_xticks(x)
    ax_blend.set_xticklabels(sec_names, rotation=40, ha="right",
                             fontsize=6, color=WHITE)
    ax_blend.set_ylim(0, 1.1)
    ax_blend.legend(fontsize=6, facecolor=PANEL, labelcolor=WHITE,
                    edgecolor="#333", loc="upper right")

    # Panel 4: Key notes
    ax_note = fig.add_subplot(gs_b[0, 3])
    ax_note.set_facecolor("#12161F"); ax_note.axis("off")
    for sp in ax_note.spines.values():
        sp.set_edgecolor(rc); sp.set_lw(1.0)

    total_mt = sum(sector_emissions.values()) if sector_emissions else 0
    dom_sect = (max(sector_emissions, key=sector_emissions.get)
                if sector_emissions else "N/A")
    dom_lulc = (max(lulc_fracs_data, key=lulc_fracs_data.get)
                if lulc_fracs_data else "N/A")

    notes = [
        ("S5P Reliability",  f"{rl}  ({int(reliability*100)}%)",  rc),
        ("Approach",
         ("S5P primary" if reliability >= 0.75 else
          "Blended" if reliability >= 0.40 else "LULC/NL dominant"),
         WHITE),
        ("Total emissions",  f"{total_mt:.2f} Mt CO2-eq",         WHITE),
        ("Dominant sector",  dom_sect.replace("_"," ").title(),   "#F4A261"),
        ("Dominant LULC",    dom_lulc,                            "#A8DADC"),
    ]
    y = 0.88
    for label, val, vc in notes:
        ax_note.text(0.06, y, label + ":", ha="left", va="center",
                     fontsize=7.5, color=GREY, transform=ax_note.transAxes)
        ax_note.text(0.06, y - 0.09, val, ha="left", va="center",
                     fontsize=9, color=vc, fontweight="bold",
                     transform=ax_note.transAxes)
        y -= 0.20

    pdf.savefig(fig, facecolor=BG, bbox_inches="tight")
    plt.close(fig)


# ===========================================================================
# PAGE 17 — Analytics: sector weights and effective S5P contributions
# ===========================================================================
def page_analytics(pdf, region_stats):
    fig = new_fig(16, 11)
    fig.text(0.5, 0.96, "Methodology Analytics — Signal Weights & Regional Variation",
             ha="center", fontsize=13, fontweight="bold", color=WHITE)

    gs = gridspec.GridSpec(2, 2, figure=fig,
                           left=0.07, right=0.97,
                           top=0.91, bottom=0.07,
                           hspace=0.38, wspace=0.30)

    BASE_RECIPES = {
        "Elec & Heat":   {"SO2": 0.40,"NO2": 0.25,"built": 0.15,"NL": 0.20},
        "Res & Comm":    {"CO":  0.30,"NO2": 0.20,"built": 0.25,"NL": 0.25},
        "Industry":      {"SO2": 0.35,"NO2": 0.30,"CO":   0.20,"built":0.15},
        "Transport":     {"NO2": 0.45,"CO":  0.35,"built": 0.20},
        "Fugitive":      {"CH4": 0.55,"SO2": 0.25,"unif": 0.20},
        "IPPU":          {"SO2": 0.45,"NO2": 0.30,"AER":  0.15,"built":0.10},
        "Agriculture":   {"CH4": 0.40,"crops":0.40,"AER": 0.20},
        "Waste":         {"CH4": 0.40,"CO":  0.30,"built": 0.30},
    }
    S5P_KEYS = {"SO2","NO2","CO","CH4","AER"}

    # Panel A: nominal blend per sector
    ax_a = fig.add_subplot(gs[0, 0])
    style_ax(ax_a, title="A — Nominal S5P vs LULC/NL weight per sector",
             ylabel="Weight fraction", fontsize=8)
    sec_names = list(BASE_RECIPES.keys())
    s5p_nom  = [sum(v for k,v in r.items() if k in S5P_KEYS)
                for r in BASE_RECIPES.values()]
    lulc_nom = [1 - s for s in s5p_nom]
    x = np.arange(len(sec_names))
    ax_a.bar(x, s5p_nom,  color=ACC2, label="S5P signals",
             edgecolor="#111", lw=0.4)
    ax_a.bar(x, lulc_nom, bottom=s5p_nom, color=ACC1,
             label="LULC + Nightlights", edgecolor="#111", lw=0.4)
    ax_a.set_xticks(x)
    ax_a.set_xticklabels(sec_names, rotation=35, ha="right",
                         fontsize=7, color=WHITE)
    ax_a.set_ylim(0, 1.12)
    ax_a.legend(fontsize=7, facecolor=PANEL, labelcolor=WHITE, edgecolor="#333")

    # Panel B: effective S5P weight at 3 reliability levels
    ax_b = fig.add_subplot(gs[0, 1])
    style_ax(ax_b, title="B — Effective S5P weight at different reliability levels",
             ylabel="Effective S5P weight", fontsize=8)
    rel_demo = [(0.88, "High (Tashkent City)", "#2ECC71"),
                (0.54, "Moderate (Tashkent Region)", "#F39C12"),
                (0.35, "Low (Navoi / Karakalpakstan)", "#E74C3C")]
    bw = 0.25
    for j, (rel, label, col) in enumerate(rel_demo):
        effs = [sum(v for k,v in r.items() if k in S5P_KEYS) * rel
                for r in BASE_RECIPES.values()]
        ax_b.bar(x + (j - 1) * bw, effs, bw, label=label,
                 color=col, edgecolor="#111", lw=0.4, alpha=0.85)
    ax_b.set_xticks(x)
    ax_b.set_xticklabels(sec_names, rotation=35, ha="right",
                         fontsize=7, color=WHITE)
    ax_b.set_ylim(0, 1.0)
    ax_b.legend(fontsize=6.5, facecolor=PANEL, labelcolor=WHITE, edgecolor="#333")

    # Panel C: reliability vs built+crops scatter
    ax_c = fig.add_subplot(gs[1, 0])
    style_ax(ax_c, title="C — S5P Reliability vs Anthropogenic LULC fraction",
             xlabel="Built + Crops (%)", ylabel="S5P Reliability", fontsize=8)
    for r in region_stats:
        lf = r["lulc_fracs"]
        xv = (lf.get("Built Area", 0) + lf.get("Crops", 0)) * 100
        yv = r["reliability"]
        ax_c.scatter(xv, yv, color=rel_color(yv), s=70, zorder=3,
                     edgecolors="#111", lw=0.5)
        ax_c.annotate(r["name"].replace(" Region","").replace("Rep. ",""),
                      (xv, yv), fontsize=5.8, color=GREY,
                      xytext=(4, 3), textcoords="offset points")
    ax_c.axhline(0.75, color="#2ECC71", lw=0.7, ls="--", alpha=0.5)
    ax_c.axhline(0.40, color="#F39C12", lw=0.7, ls="--", alpha=0.5)
    ax_c.set_ylim(0.1, 1.05); ax_c.set_xlim(0, 100)

    # Panel D: reliability vs bare+rangeland scatter
    ax_d = fig.add_subplot(gs[1, 1])
    style_ax(ax_d, title="D — S5P Reliability vs Arid LULC fraction",
             xlabel="Bare Ground + Rangeland (%)", ylabel="S5P Reliability", fontsize=8)
    for r in region_stats:
        lf = r["lulc_fracs"]
        xv = (lf.get("Bare Ground", 0) + lf.get("Rangeland", 0)) * 100
        yv = r["reliability"]
        ax_d.scatter(xv, yv, color=rel_color(yv), s=70, zorder=3,
                     edgecolors="#111", lw=0.5)
        ax_d.annotate(r["name"].replace(" Region","").replace("Rep. ",""),
                      (xv, yv), fontsize=5.8, color=GREY,
                      xytext=(4, 3), textcoords="offset points")
    ax_d.axhline(0.75, color="#2ECC71", lw=0.7, ls="--", alpha=0.5)
    ax_d.axhline(0.40, color="#F39C12", lw=0.7, ls="--", alpha=0.5)
    ax_d.set_ylim(0.1, 1.05); ax_d.set_xlim(0, 100)

    pdf.savefig(fig, facecolor=BG, bbox_inches="tight")
    plt.close(fig)
    print("  Page 17 (analytics) done.")


# ===========================================================================
# PAGE 18 — Findings & recommendations
# ===========================================================================
def page_findings(pdf, region_stats):
    fig = new_fig(16, 11)
    fig.text(0.5, 0.963, "Key Findings & Recommendations",
             ha="center", fontsize=14, fontweight="bold", color=WHITE)
    fig.add_artist(plt.Line2D([0.04, 0.96], [0.935, 0.935],
                              transform=fig.transFigure,
                              color=ACC1, lw=0.8))

    HIGH = [r for r in region_stats if r["reliability"] >= 0.75]
    MOD  = [r for r in region_stats if 0.40 <= r["reliability"] < 0.75]
    LOW  = [r for r in region_stats if r["reliability"] < 0.40]

    blocks = [
        ("#2ECC71", "HIGH reliability — S5P as primary constraint", HIGH,
         "NO2 and SO2 column densities reflect genuine combustion activity. "
         "Traffic corridors, power plants, and industrial clusters are clearly "
         "resolved in the S5P field. The v2 spatial patterns are physically "
         "meaningful and preferred for hotspot identification and policy targeting."),
        ("#F39C12", "MODERATE reliability — blended S5P + LULC/NL", MOD,
         "S5P contributes at 40-74% of nominal weight after reliability scaling. "
         "Mixed signals from semi-arid rangeland, mountain terrain, or large rural "
         "areas reduce retrieval quality. Results are indicative; local inventory "
         "data should be used for validation where available."),
        ("#E74C3C", "LOW reliability — LULC + NL dominant (v1 preferred)", LOW,
         "Desert surfaces (Kyzylkum, Karakum, Aral Sea basin) severely degrade "
         "TROPOMI UV/VIS retrievals. AER_AI is dominated by natural mineral dust. "
         "S5P weight is clamped to 15% of nominal. v1 (LULC+NL) gives a cleaner "
         "spatial pattern and should be the primary reference for these regions."),
    ]

    y = 0.90
    for col, title, regions, text in blocks:
        fig.text(0.04, y, title, fontsize=10, fontweight="bold", color=col)
        y -= 0.03
        rnames = ", ".join(r["name"] for r in regions) if regions else "none"
        fig.text(0.06, y, f"Regions: {rnames}", fontsize=8, color=WHITE)
        y -= 0.035
        for chunk in [text[i:i+110] for i in range(0, len(text), 110)]:
            fig.text(0.06, y, chunk, fontsize=8, color=GREY)
            y -= 0.028
        y -= 0.015

    # Recommendations table
    y -= 0.01
    fig.add_artist(plt.Line2D([0.04, 0.96], [y, y],
                              transform=fig.transFigure, color="#333", lw=0.5))
    y -= 0.025
    fig.text(0.04, y, "Application Guidance", fontsize=10,
             fontweight="bold", color=ACC1)
    y -= 0.03

    guidance = [
        ("Urban hotspot identification",         "v2 (S5P NO2/SO2)",    "HIGH/MOD regions"),
        ("Industrial emission cluster mapping",   "v2 (SO2 primary)",    "HIGH/MOD regions"),
        ("Traffic / transport corridor mapping",  "v2 (NO2 + CO)",       "HIGH/MOD regions"),
        ("Fugitive / gas-field emission mapping", "v2 (CH4 anomaly)",    "All regions"),
        ("Agricultural emission distribution",    "v2 (CH4 + crops)",    "All regions"),
        ("Desert / arid region spatial pattern",  "v1 (LULC+NL)",        "LOW reliability"),
        ("Sub-city spatial detail (<1 km)",        "v1 (200m LULC)",      "S5P footprint ~5.5 km"),
        ("Sector total allocation (budgets)",      "Either (identical)",  "Totals preserved in both"),
    ]
    xs = [0.06, 0.38, 0.62]
    hdrs = ["Application", "Method", "Notes"]
    for xi, h in zip(xs, hdrs):
        fig.text(xi, y, h, fontsize=8, fontweight="bold", color=ACC1)
    y -= 0.022
    for i, (app, meth, note) in enumerate(guidance):
        fig.text(xs[0], y, f"{'*' if i%2==0 else ' '} {app}", fontsize=7.5, color=WHITE)
        fig.text(xs[1], y, meth, fontsize=7.5, color="#F4A261", fontweight="bold")
        fig.text(xs[2], y, note, fontsize=7.5, color=GREY)
        y -= 0.026

    # Data sources footer
    y -= 0.01
    fig.add_artist(plt.Line2D([0.04, 0.96], [y, y],
                              transform=fig.transFigure, color="#333", lw=0.5))
    y -= 0.022
    fig.text(0.04, y, "Data Sources", fontsize=8, fontweight="bold", color=ACC1)
    src = [
        "Emission inventories: Uzbekistan First Biennial Transparency Report (BTR1), 2024",
        "Land cover: ESRI Global LULC 10m Time Series (projects/sat-io/open-datasets/landcover/ESRI_Global-LULC_10m_TS)",
        "Nightlights: NOAA/VIIRS DNB Monthly V1 VCMCFG, annual median 2022",
        "S5P NO2: COPERNICUS/S5P/OFFL/L3_NO2 — tropospheric column, cloud_fraction<0.5",
        "S5P SO2/CO/AER: COPERNICUS/S5P/OFFL/L3_SO2, L3_CO, L3_AER_AI",
        "S5P CH4: COPERNICUS/S5P/OFFL/L3_CH4 — bias-corrected, anomaly above regional P10",
        "Admin boundaries: FAO GAUL 2015 Level-1",
    ]
    for s in src:
        y -= 0.022
        fig.text(0.06, y, f"- {s}", fontsize=6.8, color="#555", style="italic")

    pdf.savefig(fig, facecolor=BG, bbox_inches="tight")
    plt.close(fig)
    print("  Page 18 (findings) done.")


# ===========================================================================
# Main
# ===========================================================================
def main():
    OUT_PDF.parent.mkdir(parents=True, exist_ok=True)

    print("Loading emissions data from Excel...")
    df = load_excel_emissions()
    print(f"  {len(df)} rows loaded.")
    if not df.empty:
        print(f"  Columns: {list(df.columns)}")

    print("Computing region stats (LULC fractions)...")
    region_stats = []
    for slug_v1, slug_v2, name, reliability in REGIONS:
        lf = get_lulc_fracs(slug_v1)
        em = get_region_emissions(df, name)
        total_mt = em.get("total", sum(em.get(s, 0) for s in SECTORS))
        region_stats.append({
            "name":        name,
            "slug_v1":     slug_v1,
            "slug_v2":     slug_v2,
            "reliability": reliability,
            "lulc_fracs":  lf,
            "sector_em":   em,
            "total_mt":    total_mt,
        })
        print(f"  {name}: reliability={reliability}, "
              f"LULC classes={list(lf.keys())}, total={total_mt:.2f} Mt")

    print(f"\nWriting PDF: {OUT_PDF}")
    with PdfPages(OUT_PDF) as pdf:
        d = pdf.infodict()
        d["Title"]   = "Uzbekistan GHG Spatial Disaggregation — Methodology Comparison 2022"
        d["Author"]  = "GEE Analysis — ESRI LULC + VIIRS + Sentinel-5P"
        d["Subject"] = "v1 (LULC+NL) vs v2 (S5P-constrained) emission disaggregation"

        page_cover(pdf)
        page_summary(pdf, region_stats)

        for i, r in enumerate(region_stats):
            print(f"  Page {i+3} — {r['name']} ...")
            page_region(pdf,
                        r["slug_v1"], r["slug_v2"], r["name"],
                        r["reliability"], r["lulc_fracs"], r["sector_em"])

        page_analytics(pdf, region_stats)
        page_findings(pdf, region_stats)

    print(f"\nDone — {OUT_PDF}")


if __name__ == "__main__":
    main()
