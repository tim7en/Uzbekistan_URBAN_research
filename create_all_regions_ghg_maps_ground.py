"""
GHG Emission Intensity Maps - All Uzbekistan Regions (2022)
v3: Ground-Level Industrial Monitoring Data Constraints

Primary: Real in-situ pollutant measurements at industrial facilities
         (power plants, chemical plants, gold mining, oil/gas, cement)
Secondary: ESRI 10m LULC (structural land-use constraint)
Tertiary: VIIRS nightlights (fill for unmonitored areas)
"""
import sys, json, math
from pathlib import Path
ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.patches import Polygon as MplPolygon
from matplotlib.collections import PatchCollection
from scipy.ndimage import gaussian_filter, zoom
import pandas as pd
import openpyxl
import rasterio
from rasterio.features import geometry_mask
from rasterio.transform import from_bounds

# ---------------------------------------------------------------------------
# Paths & constants
# ---------------------------------------------------------------------------
OUT_DIR   = ROOT / "suhi_analysis_output" / "regional_ghg_maps_ground"
CACHE_DIR = ROOT / "suhi_analysis_output" / "regional_ghg_maps" / "_raster_cache"
XLSX_PATH = ROOT / "uzbekistan_emissions_by_region_2022 (1)+.xlsx"
FAC_PATH  = ROOT / "factories+extra.xlsx"
YEAR      = 2022

# ---------------------------------------------------------------------------
# uz.json
# ---------------------------------------------------------------------------
with open(ROOT / "uz.json") as _f:
    _UZ_GJ = json.load(_f)

UZ_ID_MAP = {
    "Tashkent City":        "UZTK",
    "Tashkent Region":      "UZTO",
    "Navoi Region":         "UZNW",
    "Kashkadarya Region":   "UZQA",
    "Bukhara Region":       "UZBU",
    "Samarkand Region":     "UZSA",
    "Fergana Region":       "UZFA",
    "Andijan Region":       "UZAN",
    "Namangan Region":      "UZNG",
    "Surkhandarya Region":  "UZSU",
    "Khorezm Region":       "UZXO",
    "Jizzakh Region":       "UZJI",
    "Syrdarya Region":      "UZSI",
    "Rep. Karakalpakstan":  "UZQR",
}
_UZ_FEATURES = {f["properties"]["id"]: f for f in _UZ_GJ["features"]}


def _get_uz_feature(region_name: str) -> dict:
    fid = UZ_ID_MAP.get(region_name)
    if fid and fid in _UZ_FEATURES:
        return _UZ_FEATURES[fid]
    raise KeyError(f"No uz.json feature for '{region_name}'")


def get_region_polygon_coords(region_name: str) -> list:
    feat = _get_uz_feature(region_name)
    geom = feat["geometry"]
    if geom["type"] == "Polygon":
        return [geom["coordinates"][0]]
    return [part[0] for part in geom["coordinates"]]


# ---------------------------------------------------------------------------
# Colormaps & constants
# ---------------------------------------------------------------------------
GHG_CMAP = LinearSegmentedColormap.from_list("ghg", [
    "#F7FBFF","#DEEBF7","#C6DBEF","#9ECAE1","#6BAED6",
    "#4292C6","#2171B5","#FED976","#FEB24C","#FD8D3C",
    "#FC4E2A","#E31A1C","#BD0026","#800026","#4D004B","#1A0033",
], N=512)

SECTOR_LABELS = {
    "electricity_heat":       "Electricity & Heat",
    "residential_commercial": "Residential & Commercial",
    "industry_combustion":    "Industry Combustion",
    "transport":              "Transport",
    "fugitive_emissions":     "Fugitive Emissions",
    "ippu":                   "IPPU",
    "agriculture":            "Agriculture",
    "waste":                  "Waste",
}
SECTOR_COLORS = ["#E63946","#457B9D","#F4A261","#2A9D8F",
                 "#E9C46A","#264653","#A8DADC","#F1FAEE"]

ESRI_CLASSES = {1:"Water",2:"Trees",4:"Flooded_Vegetation",5:"Crops",
                7:"Built_Area",8:"Bare_Ground",9:"Snow_Ice",10:"Clouds",11:"Rangeland"}
BUILT=7; CROPS=5; FLOODED=4; TREES=2; RANGELAND=11; BARE=8; WATER=1

ESRI_ID_PALETTE = {
    1:"#419BDF",2:"#397D49",4:"#7A87C6",5:"#E49635",
    7:"#C4281B",8:"#A59B8F",9:"#B39FE1",11:"#DFC35A",
}

# ---------------------------------------------------------------------------
# Ground recipes
# ---------------------------------------------------------------------------
GROUND_RECIPES = {
    "electricity_heat":       {"so2": 0.40, "nox": 0.30, "no2": 0.15, "lulc_built": 0.10, "nl": 0.05},
    "residential_commercial": {"co":  0.25, "no2": 0.20, "pm25":0.15, "lulc_built": 0.25, "nl": 0.15},
    "industry_combustion":    {"so2": 0.30, "nox": 0.25, "co":  0.20, "pm10":0.10, "lulc_built": 0.15},
    "transport":              {"no2": 0.40, "co":  0.30, "pm25":0.10, "lulc_built": 0.20},
    "fugitive_emissions":     {"h2s": 0.35, "so2": 0.25, "co":  0.20, "lulc_bare":  0.20},
    "ippu":                   {"so2": 0.35, "nh3": 0.25, "no2": 0.20, "nox": 0.10, "lulc_built": 0.10},
    "agriculture":            {"nh3": 0.45, "pm10":0.15, "lulc_crops": 0.40},
    "waste":                  {"co":  0.35, "pm25":0.20, "nh3": 0.10, "lulc_built": 0.35},
}

POLL_KEYS = ["so2","no2","nox","co","nh3","h2s","pm25","pm10"]

# Facility type classification rules
TYPE_RULES = [
    ("power_plant",    ["IES","TES","TES-PGU","IEM","issiqlik elektr","\u0422\u042d\u0421"]),
    ("gold_mining",    ["NGMK","KON-METALLURGIYA","GMZ"]),
    ("chemical",       ["NAVOIYAZOT","NAZOT","AMMOFOS","fosforit","MAXAM"]),
    ("oil_gas",        ["LUKOIL","LUKOJL","O'ztransgaz","NPZ","EQMS","neft","\u041b\u0423\u041a\u041e\u0419\u041b"]),
    ("cement",         ["Sement","cement","HUAXIN"]),
    ("manufacturing",  ["zavod","ZAVOD","Zavodi","TTZ","mexanika"]),
    ("air_quality",    ["AIRCONTROL","Uzhydromet"]),
    ("weather",        ["AWS"]),
]

# Display style per type
FAC_STYLE = {
    "power_plant":    {"marker": "^", "color": "red",       "size": 120, "label": "Power Plant"},
    "gold_mining":    {"marker": "*", "color": "gold",      "size": 100, "label": "Gold Mining"},
    "chemical":       {"marker": "o", "color": "orange",    "size": 80,  "label": "Chemical"},
    "oil_gas":        {"marker": "D", "color": "#8B4513",   "size": 80,  "label": "Oil & Gas"},
    "cement":         {"marker": "s", "color": "cyan",      "size": 70,  "label": "Cement"},
    "manufacturing":  {"marker": "o", "color": "yellow",    "size": 50,  "label": "Manufacturing"},
    "air_quality":    {"marker": "o", "color": "#4488FF",   "size": 40,  "label": "Air Quality"},
    "weather":        {"marker": ".", "color": "#888888",   "size": 15,  "label": "Weather"},
    "other":          {"marker": "o", "color": "yellow",    "size": 50,  "label": "Other"},
}

INDUSTRIAL_TYPES = {"power_plant","gold_mining","chemical","oil_gas","cement","manufacturing"}
LABEL_TYPES      = {"power_plant","gold_mining","chemical","oil_gas"}

REGIONS = [
    ("Tashkent City",        "Tashkent_City"),
    ("Tashkent Region",      "Tashkent_Region"),
    ("Navoi Region",         "Navoi_Region"),
    ("Kashkadarya Region",   "Kashkadarya_Region"),
    ("Bukhara Region",       "Bukhara_Region"),
    ("Samarkand Region",     "Samarkand_Region"),
    ("Fergana Region",       "Fergana_Region"),
    ("Andijan Region",       "Andijan_Region"),
    ("Namangan Region",      "Namangan_Region"),
    ("Surkhandarya Region",  "Surkhandarya_Region"),
    ("Khorezm Region",       "Khorezm_Region"),
    ("Jizzakh Region",       "Jizzakh_Region"),
    ("Syrdarya Region",      "Syrdarya_Region"),
    ("Rep. Karakalpakstan",  "Rep_Karakalpakstan"),
]

# ---------------------------------------------------------------------------
# Facility type classifier
# ---------------------------------------------------------------------------

def classify_facility(qurilma: str, name: str, location: str) -> str:
    combined = " ".join([str(qurilma or ""), str(name or ""), str(location or "")]).lower()
    for ftype, keywords in TYPE_RULES:
        for kw in keywords:
            if kw.lower() in combined:
                return ftype
    return "other"


# ---------------------------------------------------------------------------
# Load factory monitoring data
# ---------------------------------------------------------------------------

def load_factory_data(path: Path) -> list:
    """
    Load factory monitoring sheet. Returns list of per-facility dicts:
      {qurilma, name, location, type, lat, lon, so2, no2, nox, co, nh3, h2s, pm25, pm10, total_poll}
    """
    print("Loading factory monitoring data...")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]
    # Map column names to indices
    col_idx = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        hs = str(h).strip()
        col_idx[hs] = i

    # Build column index map by simplified name
    def find_col(candidates):
        for c in candidates:
            if c in col_idx:
                return col_idx[c]
        return None

    # Column indices
    ci = {
        "qurilma":  find_col(["Qurilma"]),
        "name":     find_col(["Qurilma Nomi"]),
        "location": find_col(["Joylashuv"]),
        "pm25":     find_col(["PM2.5 (mg/m\u00b3)", "PM2.5 (mg/m3)"]),
        "pm10":     find_col(["PM10 (mg/m\u00b3)", "PM10 (mg/m3)"]),
        "so2":      find_col(["SO2 (mg/m\u00b3)", "SO2 (mg/m3)"]),
        "no":       find_col(["NO (mg/m\u00b3)", "NO (mg/m3)"]),
        "no2":      find_col(["NO2 (mg/m\u00b3)", "NO2 (mg/m3)"]),
        "nox":      find_col(["NOx (mg/m\u00b3)", "NOx (mg/m3)"]),
        "co":       find_col(["CO (mg/m\u00b3)", "CO (mg/m3)"]),
        "o3":       find_col(["O3 (mg/m\u00b3)", "O3 (mg/m3)"]),
        "nh3":      find_col(["NH3 (mg/m\u00b3)", "NH3 (mg/m3)"]),
        "h2s":      find_col(["H2S (mg/m\u00b3)", "H2S (mg/m3)"]),
        "lat":      find_col(["Latitude"]),
        "lon":      find_col(["Longitude"]),
    }

    def safe_float(v):
        if v is None:
            return float("nan")
        try:
            return float(v)
        except (TypeError, ValueError):
            return float("nan")

    # Accumulate per facility
    fac_accum = {}  # qurilma -> {poll_key: [values], lat: [], lon: [], name, location}

    for row in rows[1:]:
        q = row[ci["qurilma"]] if ci["qurilma"] is not None else None
        if not q:
            continue
        q = str(q).strip()

        lat = safe_float(row[ci["lat"]] if ci["lat"] is not None else None)
        lon = safe_float(row[ci["lon"]] if ci["lon"] is not None else None)

        nm  = str(row[ci["name"]]     if ci["name"]     is not None else "") or ""
        loc = str(row[ci["location"]] if ci["location"] is not None else "") or ""

        if q not in fac_accum:
            fac_accum[q] = {pk: [] for pk in POLL_KEYS}
            fac_accum[q]["lat"] = []
            fac_accum[q]["lon"] = []
            fac_accum[q]["name"] = nm
            fac_accum[q]["location"] = loc

        if math.isfinite(lat) and math.isfinite(lon):
            fac_accum[q]["lat"].append(lat)
            fac_accum[q]["lon"].append(lon)

        for pk in POLL_KEYS:
            v = safe_float(row[ci[pk]] if ci[pk] is not None else None)
            if math.isfinite(v):
                fac_accum[q][pk].append(v)

    # Build facility list
    facilities = []
    for q, acc in fac_accum.items():
        lats = acc["lat"]
        lons = acc["lon"]
        if not lats or not lons:
            continue
        lat = float(np.mean(lats))
        lon = float(np.mean(lons))
        if not (math.isfinite(lat) and math.isfinite(lon)):
            continue

        fac = {
            "qurilma": q,
            "name": acc["name"],
            "location": acc["location"],
            "type": classify_facility(q, acc["name"], acc["location"]),
            "lat": lat,
            "lon": lon,
        }
        total_poll = 0.0
        for pk in POLL_KEYS:
            vals = acc[pk]
            mean_v = float(np.mean(vals)) if vals else 0.0
            fac[pk] = mean_v
            total_poll += mean_v
        fac["total_poll"] = total_poll
        facilities.append(fac)

    print(f"  Loaded {len(facilities)} facilities with valid coordinates")
    return facilities


# ---------------------------------------------------------------------------
# Ground reliability
# ---------------------------------------------------------------------------

def ground_reliability(facilities_in_region: list) -> float:
    n_industrial = sum(1 for f in facilities_in_region if f["type"] in INDUSTRIAL_TYPES)
    score = 0.10 + 0.80 * (1 - math.exp(-n_industrial / 5.0))
    return min(score, 0.95)


# ---------------------------------------------------------------------------
# Gaussian kernel spreading
# ---------------------------------------------------------------------------

def spread_facilities(facilities: list, poll_key: str,
                      grid_shape: tuple, transform,
                      full_mask: np.ndarray, sigma_px: int = 15) -> np.ndarray:
    """Create a 2D kernel-spread field from facility point measurements."""
    field = np.zeros(grid_shape, dtype=float)
    for fac in facilities:
        val = fac.get(poll_key, 0)
        if not (val > 0 and math.isfinite(val)):
            continue
        lat, lon = fac["lat"], fac["lon"]
        col, row = ~transform * (lon, lat)
        col, row = int(round(col)), int(round(row))
        if not (0 <= row < grid_shape[0] and 0 <= col < grid_shape[1]):
            continue
        field[row, col] += val
    if field.max() > 0:
        field = gaussian_filter(field, sigma=sigma_px)
    field[~full_mask] = 0
    return field


# ---------------------------------------------------------------------------
# Load rasters
# ---------------------------------------------------------------------------

def load_raster(path: Path):
    with rasterio.open(path) as src:
        data = src.read(1, masked=True)
        transform = src.transform
        bounds = src.bounds
    return data, transform, bounds


def make_polygon_mask(geom_json: dict, transform, shape: tuple) -> np.ndarray:
    outside = geometry_mask([geom_json], transform=transform,
                            out_shape=shape, invert=False)
    return ~outside


def resample_to(arr: np.ma.MaskedArray, target_shape: tuple) -> np.ma.MaskedArray:
    zy = target_shape[0] / arr.shape[0]
    zx = target_shape[1] / arr.shape[1]
    data   = zoom(arr.filled(0).astype(float), (zy, zx), order=1)
    msk    = arr.mask
    mask_z = (np.zeros(target_shape, bool) if (np.isscalar(msk) or msk.shape == ())
              else zoom(msk.astype(float), (zy, zx), order=0) > 0.5)
    return np.ma.array(data, mask=mask_z)


def normalise(arr: np.ndarray, full_mask: np.ndarray, clip_pct: float = 99.0) -> np.ndarray:
    vals = arr[full_mask & np.isfinite(arr) & (arr > 0)]
    if vals.size == 0:
        return np.zeros_like(arr)
    vmax = np.percentile(vals, clip_pct)
    vmin = vals.min()
    if vmax <= vmin:
        return np.zeros_like(arr)
    out = np.clip((arr - vmin) / (vmax - vmin), 0, 1)
    out[~full_mask] = 0
    return out


# ---------------------------------------------------------------------------
# Emission grid computation
# ---------------------------------------------------------------------------

def compute_emission_grid(region_name: str, facilities_in_region: list,
                          lulc, nl_res, refined_Mt: dict,
                          pixel_area_km2: float, poly_mask: np.ndarray,
                          lulc_transform) -> np.ndarray:
    """Build the blended ground+LULC emission intensity grid."""
    lulc_data = lulc.filled(0).astype(int)
    nodata_mask = (~lulc.mask) if (hasattr(lulc, "mask") and lulc.mask.ndim > 0) \
                  else np.ones(lulc.shape, bool)
    full_mask = nodata_mask & poly_mask

    # NL normalised
    nl_data  = nl_res.filled(0).astype(float)
    nl_maxv  = np.percentile(nl_data[nl_data > 0], 98) if (nl_data > 0).any() else 1.0
    nl_norm  = np.clip(nl_data / max(nl_maxv, 1e-9), 0, 1)
    nl_norm[~full_mask] = 0

    # LULC fraction layers
    built_frac  = normalise((lulc_data == BUILT).astype(float),  full_mask)
    crops_frac  = normalise((lulc_data == CROPS).astype(float),  full_mask)
    bare_frac   = normalise((lulc_data == BARE).astype(float),   full_mask)

    # Ground reliability
    rel = ground_reliability(facilities_in_region)

    # Sigma for different types
    def sigma_for(fac):
        return 5 if fac["type"] in ("weather","air_quality") else 15

    # Build spread fields for each pollutant
    spread = {}
    for pk in POLL_KEYS:
        # Industrial point sources at high sigma
        ind_facs = [f for f in facilities_in_region if f["type"] in INDUSTRIAL_TYPES]
        bg_facs  = [f for f in facilities_in_region if f["type"] not in INDUSTRIAL_TYPES]
        field_ind = spread_facilities(ind_facs, pk, lulc.shape, lulc_transform, full_mask, sigma_px=15)
        field_bg  = spread_facilities(bg_facs,  pk, lulc.shape, lulc_transform, full_mask, sigma_px=5)
        field = field_ind + field_bg * 0.3  # background fill weighted down
        spread[pk] = normalise(field, full_mask)

    total_grid = np.zeros(lulc.shape, dtype=float)

    for sector, recipe in GROUND_RECIPES.items():
        mt = refined_Mt.get(sector) or 0.0
        if mt == 0:
            continue

        # Ground proxy
        proxy_ground = np.zeros(lulc.shape, dtype=float)
        for component, weight in recipe.items():
            if component in spread:
                proxy_ground += weight * spread[component]
            elif component == "lulc_built":
                proxy_ground += weight * built_frac
            elif component == "lulc_crops":
                proxy_ground += weight * crops_frac
            elif component == "lulc_bare":
                proxy_ground += weight * bare_frac
            elif component == "nl":
                proxy_ground += weight * nl_norm

        # LULC-only fallback proxy (simple)
        built_w = sum(w for c, w in recipe.items() if "built" in c)
        crops_w = sum(w for c, w in recipe.items() if "crop" in c)
        bare_w  = sum(w for c, w in recipe.items() if "bare"  in c)
        nl_w    = sum(w for c, w in recipe.items() if c == "nl")
        proxy_lulc = (built_w * built_frac + crops_w * crops_frac +
                      bare_w * bare_frac + nl_w * nl_norm)
        # Fill remainder with built+nl
        remaining = max(0, 1.0 - built_w - crops_w - bare_w - nl_w)
        proxy_lulc += remaining * (0.5 * built_frac + 0.5 * nl_norm)

        # Blend
        proxy = rel * proxy_ground + (1 - rel) * proxy_lulc
        proxy[~full_mask] = 0

        s = proxy[full_mask].sum()
        if s > 0:
            total_grid += (proxy / s) * mt * 1e6 / pixel_area_km2

    # Gaussian smoothing within polygon
    total_sm = gaussian_filter(total_grid, sigma=4)
    total_sm[~poly_mask] = 0.0
    return total_sm


# ---------------------------------------------------------------------------
# Load emissions Excel
# ---------------------------------------------------------------------------

def load_emissions_excel(xlsx_path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Emissions by Region"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header row: contains "Region" and "Electric"
    header_idx = None
    for i, r in enumerate(rows):
        vals = [str(v) for v in r if v is not None]
        if any("Region" in v for v in vals) and any("Electric" in v for v in vals):
            header_idx = i
            break
    if header_idx is None:
        header_idx = next(i for i, r in enumerate(rows)
                          if any(v is not None and "Region" in str(v) for v in r))

    data_rows = rows[header_idx + 1:]
    n_cols = max(len(r) for r in data_rows)
    df = pd.DataFrame([list(r) + [None] * (n_cols - len(r)) for r in data_rows],
                      columns=range(n_cols))
    col_names = ["region","electricity_heat","residential_commercial",
                 "industry_combustion","transport","fugitive_emissions",
                 "ippu","agriculture","waste","lulucf","total",
                 "national_share","net_total"]
    df = df.rename(columns={i: col_names[i] for i in range(min(len(col_names), n_cols))})
    df["region"] = df["region"].astype(str).str.strip()
    for p in ["nan","None","Note","Source","SUM OF","NATIONAL","Allocation"]:
        df = df[~df["region"].str.startswith(p)]
    for col in col_names[1:]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


# ---------------------------------------------------------------------------
# LULC composition
# ---------------------------------------------------------------------------

def lulc_composition(lulc) -> dict:
    data = lulc.filled(0).astype(int)
    valid = data[data > 0]
    if valid.size == 0:
        return {}
    total = valid.size
    result = {}
    for cid, name in ESRI_CLASSES.items():
        count = (data == cid).sum()
        if count > 0:
            result[name] = count / total
    return result


# ---------------------------------------------------------------------------
# Render map
# ---------------------------------------------------------------------------

def draw_map(region_name: str, slug: str, emission_grid: np.ndarray,
             lulc, nl_data, lulc_transform, lulc_bounds,
             refined_Mt: dict, original_Mt: dict,
             facilities_in_region: list, all_facilities: list,
             rel_score: float, out_path: Path,
             poly_rings: list, poly_mask: np.ndarray):

    rows_g, cols_g = emission_grid.shape
    west  = lulc_bounds.left
    east  = lulc_bounds.right
    south = lulc_bounds.bottom
    north = lulc_bounds.top
    extent = [west, east, south, north]

    pos_vals = emission_grid[emission_grid > 0]
    if pos_vals.size == 0:
        print(f"  WARNING: no positive emission values for {region_name}")
        return

    display = np.log10(np.maximum(emission_grid, 1.0))
    log_min = float(np.log10(max(np.percentile(pos_vals, 2), 1.0)))
    log_max = float(np.log10(np.percentile(pos_vals, 99.5)))
    if log_max <= log_min:
        log_max = log_min + 1

    # -----------------------------------------------------------------------
    # Figure layout: 3 rows
    #   Row 0 (70%) : main map
    #   Row 1 (5%)  : colorbar
    #   Row 2 (25%) : 4 panels
    # -----------------------------------------------------------------------
    fig = plt.figure(figsize=(16, 12))
    fig.patch.set_facecolor("#0E1117")

    gs = gridspec.GridSpec(
        3, 4, figure=fig,
        height_ratios=[14, 1, 5],
        hspace=0.40, wspace=0.35,
        top=0.93, bottom=0.04, left=0.06, right=0.97,
    )

    ax_map  = fig.add_subplot(gs[0, :])
    ax_cbar = fig.add_subplot(gs[1, :])
    ax_bars = fig.add_subplot(gs[2, 0])
    ax_pie  = fig.add_subplot(gs[2, 1])
    ax_poll = fig.add_subplot(gs[2, 2])
    ax_info = fig.add_subplot(gs[2, 3])

    # ---- LULC faint background ----
    import matplotlib.colors as mcolors
    lulc_int = lulc.filled(0).astype(int)
    lulc_rgba = np.zeros((*lulc.shape, 4), dtype=float)
    for cid, hex_col in ESRI_ID_PALETTE.items():
        rgba = mcolors.to_rgba(hex_col, 0.18)
        px_mask = (lulc_int == cid) & poly_mask
        for c in range(4):
            lulc_rgba[px_mask, c] = rgba[c]
    ax_map.imshow(lulc_rgba, extent=extent, origin="upper", aspect="auto",
                  interpolation="nearest")

    # ---- Masked outside (dark) ----
    outside_rgba = np.zeros((*lulc.shape, 4), dtype=float)
    outside_rgba[~poly_mask, :] = [0.05, 0.06, 0.09, 0.85]
    ax_map.imshow(outside_rgba, extent=extent, origin="upper", aspect="auto",
                  interpolation="nearest")

    # ---- GHG intensity ----
    em_display = display.copy()
    em_display[~poly_mask] = np.nan
    img = ax_map.imshow(em_display, extent=extent, origin="upper",
                        cmap=GHG_CMAP, vmin=log_min, vmax=log_max,
                        interpolation="bilinear", aspect="auto", alpha=0.93)

    # ---- Region boundary ----
    if poly_rings:
        for ring in poly_rings:
            xs = [c[0] for c in ring] + [ring[0][0]]
            ys = [c[1] for c in ring] + [ring[0][1]]
            ax_map.plot(xs, ys, color="white", lw=3, alpha=0.4, zorder=5)
            ax_map.plot(xs, ys, color="white", lw=1.5, alpha=1.0, zorder=6)

    # ---- Facility dots ----
    for fac in facilities_in_region:
        st = FAC_STYLE.get(fac["type"], FAC_STYLE["other"])
        sz = st["size"]
        # Scale size by total_poll (log-weighted)
        tp = fac.get("total_poll", 0)
        if tp > 0:
            sz = sz * (1.0 + 0.5 * math.log1p(tp))
        sz = min(sz, sz * 3)
        ax_map.scatter(fac["lon"], fac["lat"],
                       marker=st["marker"], color=st["color"],
                       s=sz, zorder=10, edgecolors="white", linewidths=0.5)
        # Label for key industrial types
        if fac["type"] in LABEL_TYPES:
            label = fac["name"][:20] if fac["name"] else fac["qurilma"][:20]
            ax_map.text(fac["lon"] + 0.02, fac["lat"] + 0.02, label,
                        fontsize=5.5, color="white", zorder=11,
                        bbox=dict(facecolor="#0E1117", alpha=0.4,
                                  edgecolor="none", pad=0.5))

    # ---- Map style ----
    total_mt = sum(v for k, v in refined_Mt.items()
                   if k != "lulucf" and v is not None and v > 0)
    ax_map.set_title(
        f"{region_name}  -  Total: {total_mt:.2f} Mt CO2-eq (excl. LULUCF)",
        fontsize=10, fontweight="bold", color="white", pad=6,
    )
    ax_map.set_xlabel("Longitude", fontsize=8, color="white")
    ax_map.set_ylabel("Latitude",  fontsize=8, color="white")
    ax_map.tick_params(colors="white", labelsize=7)
    for sp in ax_map.spines.values():
        sp.set_edgecolor("#444")
    ax_map.set_facecolor("#0E1117")

    # v3 annotation
    ax_map.text(0.01, 0.97, "v3: Ground Monitoring",
                transform=ax_map.transAxes, fontsize=8, color="#AAFFAA",
                fontweight="bold", va="top",
                bbox=dict(facecolor="#0E1117", alpha=0.6, edgecolor="none", pad=2))

    # Facility type legend in map
    shown_types = set(f["type"] for f in facilities_in_region)
    legend_handles = []
    for ftype in ["power_plant","gold_mining","chemical","oil_gas","cement",
                  "manufacturing","air_quality","weather","other"]:
        if ftype in shown_types:
            st = FAC_STYLE[ftype]
            legend_handles.append(
                plt.scatter([], [], marker=st["marker"], color=st["color"],
                            s=40, label=st["label"], edgecolors="white", linewidths=0.4)
            )
    if legend_handles:
        leg = ax_map.legend(handles=legend_handles, loc="lower right",
                            fontsize=6, facecolor="#16191F", labelcolor="white",
                            edgecolor="#444", framealpha=0.85, ncol=2)

    # ---- Colorbar ----
    cb = fig.colorbar(img, cax=ax_cbar, orientation="horizontal", extend="max")
    cb.set_label("GHG Emission Intensity  (t CO2-eq / km2)",
                 fontsize=8, color="white", labelpad=3)
    tick_vals = [1, 5, 10, 50, 100, 500, 1000, 5000, 10000, 25000]
    t_log = [np.log10(v) for v in tick_vals if log_min <= np.log10(v) <= log_max]
    t_lbl = [f"{v:,}" for v in tick_vals if log_min <= np.log10(v) <= log_max]
    if t_log:
        cb.set_ticks(t_log)
        cb.set_ticklabels(t_lbl)
    cb.ax.tick_params(colors="white", labelsize=7)
    cb.outline.set_edgecolor("#777")
    ax_cbar.set_facecolor("#0E1117")

    # ---- Panel 1: Sector emissions bar chart ----
    ax_bars.set_facecolor("#16191F")
    sectors = list(SECTOR_LABELS.keys())
    x = np.arange(len(sectors))
    orig_vals = [float(original_Mt.get(s) or 0) for s in sectors]
    bars = ax_bars.bar(x, orig_vals, 0.55,
                       color=SECTOR_COLORS[:len(sectors)],
                       edgecolor="#0E1117", lw=0.5, zorder=3)
    max_v = max(orig_vals) if any(v > 0 for v in orig_vals) else 1
    for bar, v in zip(bars, orig_vals):
        if v > 0:
            ax_bars.text(bar.get_x() + bar.get_width() / 2,
                         bar.get_height() + max_v * 0.02,
                         f"{v:.2f}", ha="center", va="bottom",
                         fontsize=6, color="white", fontweight="bold")
    ax_bars.set_xticks(x)
    ax_bars.set_xticklabels([SECTOR_LABELS[s] for s in sectors],
                            rotation=35, ha="right", fontsize=6, color="white")
    ax_bars.set_ylabel("Mt CO2-eq", fontsize=7, color="white")
    ax_bars.set_title("Sector Emissions (BTR1)",
                      fontsize=8, fontweight="bold", color="white", pad=4)
    ax_bars.tick_params(colors="white", labelsize=7)
    ax_bars.set_ylim(0, max_v * 1.3)
    ax_bars.spines["top"].set_visible(False)
    ax_bars.spines["right"].set_visible(False)
    for sp in ax_bars.spines.values():
        sp.set_edgecolor("#444")
    ax_bars.yaxis.grid(True, alpha=0.15, lw=0.5, color="white")
    ax_bars.set_axisbelow(True)

    # ---- Panel 2: Facility type pie ----
    ax_pie.set_facecolor("#16191F")
    ax_pie.set_title("Monitoring Stations by Type",
                     fontsize=8, fontweight="bold", color="white", pad=4)
    type_counts = {}
    for f in facilities_in_region:
        type_counts[f["type"]] = type_counts.get(f["type"], 0) + 1
    if type_counts:
        pie_labels = list(type_counts.keys())
        pie_sizes  = [type_counts[t] for t in pie_labels]
        pie_colors = [FAC_STYLE.get(t, FAC_STYLE["other"])["color"] for t in pie_labels]
        wedges, texts, autotexts = ax_pie.pie(
            pie_sizes, colors=pie_colors,
            autopct=lambda p: f"{p:.0f}%" if p > 6 else "",
            startangle=90,
            wedgeprops={"edgecolor": "#16191F", "linewidth": 0.8},
            pctdistance=0.75,
            textprops={"fontsize": 6, "color": "white"},
        )
        for t in autotexts:
            t.set_fontsize(5.5)
        pie_patches = [
            mpatches.Patch(
                color=FAC_STYLE.get(pie_labels[i], FAC_STYLE["other"])["color"],
                label=f"{FAC_STYLE.get(pie_labels[i], FAC_STYLE['other'])['label']}: {pie_sizes[i]}"
            )
            for i in range(len(pie_labels))
        ]
        ax_pie.legend(handles=pie_patches, loc="lower center",
                      bbox_to_anchor=(0.5, -0.35), fontsize=5.5, ncol=2,
                      facecolor="#16191F", labelcolor="white",
                      edgecolor="#444", framealpha=0.8)
    else:
        ax_pie.text(0.5, 0.5, "No facilities", ha="center", va="center",
                    transform=ax_pie.transAxes, color="white", fontsize=8)
        ax_pie.axis("off")

    # ---- Panel 3: Top pollutant bar chart ----
    ax_poll.set_facecolor("#16191F")
    ax_poll.set_title("Mean Industrial Pollutant Conc.",
                      fontsize=8, fontweight="bold", color="white", pad=4)
    ind_facs_reg = [f for f in facilities_in_region if f["type"] in INDUSTRIAL_TYPES]
    poll_display = ["so2","no2","nox","co","nh3","h2s"]
    poll_labels  = ["SO2","NO2","NOx","CO","NH3","H2S"]
    poll_colors  = ["#E63946","#F4A261","#FED976","#A8DADC","#457B9D","#264653"]
    if ind_facs_reg:
        poll_means = []
        for pk in poll_display:
            vals = [f[pk] for f in ind_facs_reg if f[pk] > 0]
            poll_means.append(float(np.mean(vals)) if vals else 0.0)
        px = np.arange(len(poll_display))
        pbars = ax_poll.bar(px, poll_means, 0.6,
                            color=poll_colors, edgecolor="#0E1117", lw=0.5)
        ax_poll.set_xticks(px)
        ax_poll.set_xticklabels(poll_labels, fontsize=7, color="white")
        ax_poll.set_ylabel("mg/m3", fontsize=7, color="white")
        pmax = max(poll_means) if any(v > 0 for v in poll_means) else 1
        ax_poll.set_ylim(0, pmax * 1.35)
        for bar, v in zip(pbars, poll_means):
            if v > 0:
                ax_poll.text(bar.get_x() + bar.get_width() / 2,
                             bar.get_height() + pmax * 0.02,
                             f"{v:.3f}", ha="center", va="bottom",
                             fontsize=5.5, color="white")
    else:
        ax_poll.text(0.5, 0.5, "No industrial\nfacilities",
                     ha="center", va="center",
                     transform=ax_poll.transAxes, color="white", fontsize=8)
        ax_poll.axis("off")
    ax_poll.tick_params(colors="white", labelsize=7)
    ax_poll.spines["top"].set_visible(False)
    ax_poll.spines["right"].set_visible(False)
    for sp in ax_poll.spines.values():
        sp.set_edgecolor("#444")
    ax_poll.yaxis.grid(True, alpha=0.15, lw=0.5, color="white")
    ax_poll.set_axisbelow(True)

    # ---- Panel 4: Info box ----
    ax_info.set_facecolor("#16191F")
    ax_info.axis("off")
    for sp in ax_info.spines.values():
        sp.set_edgecolor("#444")

    if rel_score >= 0.65:
        rel_label = "HIGH"
        rel_color = "#44FF88"
    elif rel_score >= 0.35:
        rel_label = "MOD"
        rel_color = "#FFD700"
    else:
        rel_label = "LOW"
        rel_color = "#FF6644"

    dominant_sector = max(
        [(s, float(original_Mt.get(s) or 0)) for s in SECTOR_LABELS],
        key=lambda x: x[1]
    )
    dom_label = SECTOR_LABELS.get(dominant_sector[0], dominant_sector[0])

    total_mt_orig = sum(float(v) for k, v in original_Mt.items()
                        if k != "lulucf" and v is not None and not math.isnan(float(v)) and v > 0)

    n_all = len(facilities_in_region)
    n_ind = sum(1 for f in facilities_in_region if f["type"] in INDUSTRIAL_TYPES)

    # info_lines: (text, color, fontsize, fontweight, fontstyle)
    info_lines = [
        (region_name,                               "white",    8.5, "bold",   "normal"),
        ("",                                        "white",    4,   "normal", "normal"),
        (f"Ground Reliability: {rel_label}  ({rel_score:.2f})", rel_color, 8, "bold", "normal"),
        (f"Total GHG: {total_mt_orig:.2f} Mt CO2-eq", "white", 7.5, "normal", "normal"),
        (f"Dominant: {dom_label}",                  "white",    7,   "normal", "normal"),
        ("",                                        "white",    3,   "normal", "normal"),
        (f"Monitoring stations: {n_all}",           "#CCCCCC",  7,   "normal", "normal"),
        (f"Industrial monitors: {n_ind}",           "#CCCCCC",  7,   "normal", "normal"),
        ("",                                        "white",    3,   "normal", "normal"),
        ("v3: Ground-Level",                        "#AAFFAA",  7,   "normal", "italic"),
        ("Industrial Monitoring",                   "#AAFFAA",  7,   "normal", "italic"),
    ]

    y_pos = 0.96
    for text, color, size, weight, style in info_lines:
        if not text:
            y_pos -= 0.03
            continue
        ax_info.text(0.05, y_pos, text, transform=ax_info.transAxes,
                     fontsize=size, color=color, fontweight=weight,
                     fontstyle=style, va="top")
        y_pos -= size * 0.013 + 0.02

    # ---- Overall title ----
    fig.text(0.5, 0.965,
             f"GHG Emissions - {region_name} (2022)  [v3: Ground Monitoring]",
             ha="center", fontsize=13, fontweight="bold", color="white")
    fig.text(0.5, 0.945,
             "Spatial disaggregation: Ground industrial monitoring data + ESRI 10m LULC + VIIRS Nightlights  |  BTR1 Uzbekistan",
             ha="center", fontsize=7.5, color="#AAAAAA")

    fig.text(0.5, 0.005,
             "Sources: Industrial monitoring stations (225+ sites); ESRI Global LULC 10m; NOAA/VIIRS DNB; BTR1 Uzbekistan to UNFCCC (2024).",
             ha="center", fontsize=5.5, color="#666", style="italic")

    fig.savefig(out_path, dpi=180, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f"  Saved: {out_path.name}")


# ---------------------------------------------------------------------------
# Region geometry helpers
# ---------------------------------------------------------------------------

def facilities_in_region(facilities: list, region_name: str) -> list:
    """Return facilities whose coordinates fall inside the region polygon."""
    from shapely.geometry import Point, shape
    feat = _get_uz_feature(region_name)
    poly = shape(feat["geometry"])
    result = []
    for f in facilities:
        pt = Point(f["lon"], f["lat"])
        if pt.within(poly):
            result.append(f)
    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Load factory data
    all_facilities = load_factory_data(FAC_PATH)

    # Load emissions
    print("Loading BTR1 emissions data...")
    df_em = load_emissions_excel(XLSX_PATH)
    print(f"  Regions in Excel: {df_em['region'].tolist()}")

    def match_region_row(excel_name: str, region_name: str) -> bool:
        en = excel_name.lower().strip()
        rn = region_name.lower().strip()
        return en in rn or rn in en

    processed = []
    skipped   = []
    sectors_all = ["electricity_heat","residential_commercial","industry_combustion",
                   "transport","fugitive_emissions","ippu","agriculture","waste","lulucf"]

    for region_name, slug in REGIONS:
        print(f"\n{'='*60}")
        print(f"Processing: {region_name}")
        print(f"{'='*60}")

        out_path = OUT_DIR / f"ghg_ground_{slug}_{YEAR}.png"

        # Match Excel row
        matched = df_em[df_em["region"].apply(lambda r: match_region_row(r, region_name))]
        if matched.empty:
            print(f"  WARNING: No Excel row matched for '{region_name}' - skipping")
            skipped.append(region_name)
            continue
        row = matched.iloc[0]

        original_Mt = {s: (float(row[s]) if s in row.index and pd.notna(row[s]) else 0.0)
                       for s in sectors_all}

        # Facilities in region
        try:
            reg_facs = facilities_in_region(all_facilities, region_name)
        except Exception as e:
            print(f"  WARNING: facility point-in-polygon failed: {e}")
            reg_facs = []

        # Type counts
        type_counts = {}
        for f in reg_facs:
            type_counts[f["type"]] = type_counts.get(f["type"], 0) + 1
        print(f"  Facilities: {len(reg_facs)} total | " +
              " | ".join(f"{t}:{n}" for t, n in sorted(type_counts.items())))

        rel = ground_reliability(reg_facs)
        print(f"  Ground reliability score: {rel:.3f}")

        # Load LULC / NL from cache
        rdir = CACHE_DIR / slug
        lulc_path = rdir / f"lulc_{slug}_{YEAR}.tif"
        nl_path   = rdir / f"nl_{slug}_{YEAR}.tif"

        if not lulc_path.exists():
            print(f"  WARNING: LULC cache missing at {lulc_path} - skipping")
            skipped.append(region_name)
            continue
        if not nl_path.exists():
            print(f"  WARNING: NL cache missing at {nl_path} - skipping")
            skipped.append(region_name)
            continue

        try:
            lulc, lulc_transform, lulc_bounds = load_raster(lulc_path)
            nl,   nl_transform,   nl_bounds   = load_raster(nl_path)
        except Exception as e:
            print(f"  ERROR loading rasters: {e} - skipping")
            skipped.append(region_name)
            continue

        nl_res = resample_to(nl, lulc.shape)

        pixel_deg      = abs(lulc_transform.a)
        pixel_km       = pixel_deg * 111.0
        pixel_area_km2 = pixel_km ** 2

        # Polygon mask
        try:
            feat_geom = _get_uz_feature(region_name)["geometry"]
            poly_mask = make_polygon_mask(feat_geom, lulc_transform, lulc.shape)
            print(f"  Polygon mask: {poly_mask.sum():,} pixels inside region")
        except Exception as e:
            print(f"  WARNING: polygon mask failed ({e}), using full bbox")
            poly_mask = np.ones(lulc.shape, dtype=bool)

        # Simple Mt refinement (same approach as v1)
        lulc_comp_frac = lulc_composition(lulc)
        built_frac_s   = lulc_comp_frac.get("Built_Area", 0)
        crops_frac_s   = lulc_comp_frac.get("Crops", 0)
        flooded_frac_s = lulc_comp_frac.get("Flooded_Vegetation", 0)
        trees_frac_s   = lulc_comp_frac.get("Trees", 0)
        bare_frac_s    = lulc_comp_frac.get("Bare_Ground", 0)
        rangeland_frac_s = lulc_comp_frac.get("Rangeland", 0)

        nl_data_raw = nl_res.filled(0).astype(float)
        nl_max_r    = np.percentile(nl_data_raw[nl_data_raw > 0], 98) \
                      if (nl_data_raw > 0).any() else 1.0
        nl_score    = float(np.mean(nl_data_raw[nl_data_raw > 0]) / nl_max_r) \
                      if (nl_data_raw > 0).any() else 0.0

        def blend(prior, lulc_c, lw, nw, pw):
            adj = lw * lulc_c + nw * nl_score + pw * 1.0
            return float(prior) * adj

        refined_Mt = {
            "electricity_heat":       blend(original_Mt["electricity_heat"],
                                            built_frac_s, 0.20, 0.50, 0.30),
            "residential_commercial": blend(original_Mt["residential_commercial"],
                                            built_frac_s, 0.55, 0.20, 0.25),
            "industry_combustion":    blend(original_Mt["industry_combustion"],
                                            0.75*built_frac_s+0.25*bare_frac_s, 0.20, 0.50, 0.30),
            "transport":              blend(original_Mt["transport"],
                                            built_frac_s, 0.35, 0.35, 0.30),
            "fugitive_emissions":     float(original_Mt["fugitive_emissions"]) * (0.25*nl_score + 0.75),
            "ippu":                   blend(original_Mt["ippu"],
                                            0.5*built_frac_s+0.5*bare_frac_s, 0.10, 0.50, 0.40),
            "agriculture":            blend(original_Mt["agriculture"],
                                            0.7*crops_frac_s+0.3*flooded_frac_s, 0.65, 0.00, 0.35),
            "waste":                  blend(original_Mt["waste"], built_frac_s, 0.45, 0.00, 0.55),
            "lulucf":                 blend(original_Mt["lulucf"],
                                            trees_frac_s+0.1*rangeland_frac_s, 0.65, 0.00, 0.35),
        }

        # Build emission grid
        print(f"  Building ground emission grid ({lulc.shape[0]}x{lulc.shape[1]} px)...")
        try:
            grid = compute_emission_grid(
                region_name, reg_facs, lulc, nl_res, refined_Mt,
                pixel_area_km2, poly_mask, lulc_transform
            )
        except Exception as e:
            import traceback
            print(f"  ERROR building grid: {e}")
            traceback.print_exc()
            skipped.append(region_name)
            continue

        pos = grid[grid > 0]
        if pos.size > 0:
            print(f"  Grid range: {pos.min():.1f} - {pos.max():.1f} t/km2  (mean {pos.mean():.1f})")

        # Render map
        print(f"  Rendering map...")
        try:
            poly_rings = get_region_polygon_coords(region_name)
        except Exception:
            poly_rings = []

        try:
            draw_map(
                region_name, slug, grid, lulc, nl_data_raw, lulc_transform, lulc_bounds,
                refined_Mt, original_Mt, reg_facs, all_facilities,
                rel, out_path, poly_rings, poly_mask,
            )
            processed.append(region_name)
        except Exception as e:
            import traceback
            print(f"  ERROR rendering: {e}")
            traceback.print_exc()
            skipped.append(region_name)

    # Summary
    print(f"\n{'='*60}")
    print(f"Done.  Processed: {len(processed)}  Skipped: {len(skipped)}")
    print(f"Maps created: {len(processed)}")
    for r in processed:
        slug = dict(REGIONS)[r]
        print(f"  - ghg_ground_{slug}_{YEAR}.png")
    if skipped:
        print(f"Skipped: {skipped}")
    print(f"Output dir: {OUT_DIR}")


if __name__ == "__main__":
    main()
